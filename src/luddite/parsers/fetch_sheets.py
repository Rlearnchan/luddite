"""Fetch or normalize sheet-like inputs with redaction.

The first implementation supports local CSV/XLSX files. Google Sheet API access
will be layered in once connector/auth details are available; the redaction and
row-normalization contract is already exercised here.
"""

from __future__ import annotations

import csv
import unicodedata
from pathlib import Path
from typing import Annotated, Any

import typer
from openpyxl import load_workbook

from luddite import paths
from luddite.utils.jsonl import write_jsonl
from luddite.utils.security import (
    contains_sensitive_text,
    detect_risk_flags,
    redact_sensitive_text,
)

app = typer.Typer(no_args_is_help=False)


def _relative(path: Path) -> str:
    try:
        return str(path.relative_to(paths.REPO_ROOT))
    except ValueError:
        return str(path)


def _normalize_header(value: Any, fallback: str) -> str:
    text = str(value or "").strip()
    if not text:
        text = fallback
    text = unicodedata.normalize("NFC", text)
    text = text.replace("\n", " ").replace("\t", " ")
    return text


def _redact_cell(value: Any) -> str:
    if value is None:
        return ""
    return redact_sensitive_text(str(value))


def _row_record(
    source_file: Path,
    sheet_name: str,
    row_no: int,
    headers: list[str],
    values: list[Any],
) -> dict[str, Any]:
    raw_values = ["" if value is None else str(value) for value in values]
    redacted_values = [_redact_cell(value) for value in values]
    row = {
        header: redacted_values[index] if index < len(redacted_values) else ""
        for index, header in enumerate(headers)
    }
    credential_risk = any(contains_sensitive_text(value) for value in raw_values)
    text = " ".join(redacted_values)
    risk_flags = detect_risk_flags(text)
    if credential_risk and "needs_human_review" not in risk_flags:
        risk_flags.append("needs_human_review")

    return {
        "source_file": _relative(source_file),
        "sheet_name": sheet_name,
        "row_no": row_no,
        "row": row,
        "credential_risk": credential_risk,
        "risk_flags": sorted(set(risk_flags)),
    }


def _parse_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as source:
        reader = csv.reader(source)
        rows = list(reader)
    if not rows:
        return []
    headers = [
        _normalize_header(value, f"column_{index}")
        for index, value in enumerate(rows[0], start=1)
    ]
    return [
        _row_record(path, path.stem, row_no, headers, row)
        for row_no, row in enumerate(rows[1:], start=2)
        if any(str(value).strip() for value in row)
    ]


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [
            _normalize_header(value, f"column_{index}")
            for index, value in enumerate(rows[0], start=1)
        ]
        for row_no, row in enumerate(rows[1:], start=2):
            if not any("" if value is None else str(value).strip() for value in row):
                continue
            records.append(_row_record(path, worksheet.title, row_no, headers, list(row)))
    return records


def iter_sheet_files(source: Path) -> list[Path]:
    if source.is_file():
        return [source]
    if not source.exists():
        return []
    files: list[Path] = []
    for pattern in ("*.csv", "*.tsv", "*.xlsx"):
        files.extend(source.glob(pattern))
    return sorted(files, key=lambda path: unicodedata.normalize("NFC", path.name))


def parse_sheet_file(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _parse_csv(path)
    if suffix == ".tsv":
        with path.open(encoding="utf-8-sig") as source:
            rows = list(csv.reader(source, delimiter="\t"))
        if not rows:
            return []
        headers = [
            _normalize_header(value, f"column_{index}")
            for index, value in enumerate(rows[0], start=1)
        ]
        return [
            _row_record(path, path.stem, row_no, headers, row)
            for row_no, row in enumerate(rows[1:], start=2)
            if any(str(value).strip() for value in row)
        ]
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    raise ValueError(f"Unsupported sheet file type: {path}")


def write_redacted_csvs(records: list[dict[str, Any]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for record in records:
        grouped.setdefault((record["source_file"], record["sheet_name"]), []).append(record)

    for (source_file, sheet_name), group in grouped.items():
        first_row = group[0]["row"]
        safe_stem = Path(source_file).stem.replace("/", "_")
        output_path = output_dir / f"{safe_stem}__{sheet_name}.csv"
        with output_path.open("w", encoding="utf-8-sig", newline="") as output:
            writer = csv.DictWriter(output, fieldnames=list(first_row.keys()))
            writer.writeheader()
            for record in group:
                writer.writerow(record["row"])


def fetch_sheets(source: Path, output: Path, redacted_csv_dir: Path) -> list[dict[str, Any]]:
    if str(source).startswith(("http://", "https://")):
        raise NotImplementedError(
            "Google Sheet URL fetching is not wired yet; export CSV/XLSX locally first."
        )

    records: list[dict[str, Any]] = []
    for file_path in iter_sheet_files(source):
        records.extend(parse_sheet_file(file_path))
    write_jsonl(output, records)
    write_redacted_csvs(records, redacted_csv_dir)
    return records


@app.callback(invoke_without_command=True)
def main(
    source: Annotated[
        Path,
        typer.Option(
            "--source",
            help="Local CSV/XLSX file or directory. Google fetch comes later.",
        ),
    ] = paths.SHEETS_RAW_DIR,
    output: Annotated[
        Path,
        typer.Option("--output", help="JSONL output path."),
    ] = paths.SHEETS_PARSED_JSONL,
    redacted_csv_dir: Annotated[
        Path,
        typer.Option("--redacted-csv-dir", help="Directory for redacted CSV copies."),
    ] = paths.SHEETS_PARSED_DIR,
) -> None:
    records = fetch_sheets(source, output, redacted_csv_dir)
    typer.echo(f"Wrote {len(records)} redacted sheet rows to {output}")


if __name__ == "__main__":
    app()
